#!/usr/bin/env python3
"""
Mexico Weekly News Aggregator v2
Fetches news from configured RSS sources, filters for relevance using Claude AI,
deduplicates/merges similar stories, and outputs a formatted interactive HTML newsletter.
"""

import feedparser
import requests
from datetime import datetime, timedelta, timezone
import json
import re
import os
import time
import sys
from collections import defaultdict
from bs4 import BeautifulSoup
import anthropic

# ─────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────

TODAY = datetime.now(timezone.utc)
ONE_WEEK_AGO = TODAY - timedelta(days=7)
OUTPUT_FILE = "d:/AI Projects/Auto Weekly NL_202604/mexico_weekly_news.html"
EXCEL_FILE  = "d:/AI Projects/Auto Weekly NL_202604/source_status.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# Sources with RSS feeds to try
# URLs verified 2026-04-15
RSS_SOURCES = [
    # ── Free News Sources ──
    {
        "name": "BBC",
        # Latin America feed confirmed working: 9 articles past week
        "rss": "https://feeds.bbci.co.uk/news/world/latin_america/rss.xml",
        "rss_alt": "https://feeds.bbci.co.uk/news/rss.xml",
        "type": "free",
        "lang": "en",
    },
    {
        "name": "Reuters",
        # feeds.reuters.com is blocked; reuters.com requires auth — skipped at runtime
        "rss": "https://feeds.reuters.com/reuters/businessNews",
        "rss_alt": "https://feeds.reuters.com/reuters/worldnews",
        "type": "free",
        "lang": "en",
    },
    {
        "name": "Reforma",
        # Confirmed working: 10 articles past week
        "rss": "https://www.reforma.com/rss/portada.xml",
        "rss_alt": "https://www.reforma.com/rss",
        "type": "free",
        "lang": "es",
    },
    {
        "name": "Expansion",
        # Confirmed working: 48 articles past week
        "rss": "https://expansion.mx/rss",
        "rss_alt": "https://expansion.mx/arc/outboundfeeds/rss/",
        "type": "free",
        "lang": "es",
    },
    {
        "name": "El Universal",
        # Confirmed working via Arc feed: 100 articles past week
        "rss": "https://eluniversal.com.mx/arc/outboundfeeds/rss/?outputType=xml",
        "rss_alt": "https://www.eluniversal.com.mx/rss.xml",
        "type": "free",
        "lang": "es",
    },
    {
        "name": "El Financiero",
        # Confirmed working: 100 articles past week
        "rss": "https://www.elfinanciero.com.mx/arc/outboundfeeds/rss/?outputType=xml&size=20",
        "rss_alt": "https://www.elfinanciero.com.mx/arc/outboundfeeds/rss/",
        "type": "free",
        "lang": "es",
    },
    {
        "name": "El Economista",
        # www. version times out; no-www also returns 404 — best-effort only
        "rss": "https://www.eleconomista.com.mx/rss",
        "rss_alt": "https://eleconomista.com.mx/arc/outboundfeeds/rss/",
        "type": "free",
        "lang": "es",
    },
    {
        "name": "Bloomberg Linea",
        # Confirmed working (no-www Arc feed): 97 articles past week
        "rss": "https://bloomberglinea.com/arc/outboundfeeds/rss/?outputType=xml",
        "rss_alt": "https://bloomberglinea.com/rss/",
        "type": "free",
        "lang": "es",
    },
    {
        "name": "Mexico News Daily",
        # Confirmed working: 10 articles past week
        "rss": "https://mexiconewsdaily.com/feed/",
        "rss_alt": None,
        "type": "free",
        "lang": "en",
    },
    # ── Paid Sources (try RSS anyway) ──
    {
        "name": "Fintech Expert",
        # Substack free RSS only shows 1 preview post from 2021 — paid paywalled
        "rss": "https://fintechexpert.substack.com/feed",
        "rss_alt": "https://fintechexpert.mx/feed",
        "type": "paid",
        "lang": "es",
    },
    {
        "name": "Fintech Hub",
        # RSS accessible but last post 2026-02-27 — inactive or slow newsletter
        "rss": "https://fintechhub.com.mx/feed/",
        "rss_alt": "http://fintechhub.com.mx/feed/",
        "type": "paid",
        "lang": "es",
    },
]

# Sources we're skipping and why
SKIPPED_SOURCES = [
    {
        "name": "Ground News",
        "reason": "Paid subscription required, no public RSS/API available.",
        "suggestion": "Consider using their mobile app manually, or check if subscription includes API access.",
    },
    {
        "name": "El Lago de los Business (LinkedIn)",
        "reason": "LinkedIn actively blocks automated scraping; no RSS feed.",
        "suggestion": "Follow manually, or use Google Alerts for 'El Lago de los Business' as a proxy.",
    },
    {
        "name": "El CEO (LinkedIn)",
        "reason": "LinkedIn actively blocks automated scraping; no RSS feed.",
        "suggestion": "Follow manually, or use Google Alerts for 'El CEO Mexico' as a proxy.",
    },
    {
        "name": "Jeanette Leyva Reus (LinkedIn)",
        "reason": "LinkedIn actively blocks automated scraping; no RSS feed.",
        "suggestion": "She writes for El Financiero -- already captured via that RSS feed. Filter by her byline.",
    },
    {
        "name": "El Lago de los Business (YouTube)",
        "reason": "Channel ID is unknown (placeholder 'XXX' in config); video content requires transcription.",
        "suggestion": "Find the real channel ID from YouTube, then use: https://www.youtube.com/feeds/videos.xml?channel_id=REAL_ID",
    },
]


# ─────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────

def clean_html(text: str) -> str:
    """Strip HTML tags and normalize whitespace."""
    if not text:
        return ""
    try:
        soup = BeautifulSoup(text, "html.parser")
        return re.sub(r"\s+", " ", soup.get_text()).strip()
    except Exception:
        return re.sub(r"<[^>]+>", " ", text).strip()


def parse_date(entry):
    """Extract publication date from a feedparser entry."""
    for attr in ("published_parsed", "updated_parsed", "created_parsed"):
        val = getattr(entry, attr, None)
        if val:
            try:
                return datetime(*val[:6], tzinfo=timezone.utc)
            except Exception:
                pass
    return None


def fetch_rss(source: dict):
    """
    Fetch and parse an RSS feed.
    Returns (articles_list, error_string_or_None, detail_string_or_None).
    """
    urls_to_try = [source["rss"]]
    if source.get("rss_alt"):
        urls_to_try.append(source["rss_alt"])

    last_error = None
    for url in urls_to_try:
        try:
            resp = requests.get(url, headers=HEADERS, timeout=20, allow_redirects=True)
            resp.raise_for_status()
            feed = feedparser.parse(resp.content)

            if not feed.entries:
                last_error = f"Feed parsed but 0 entries (HTTP {resp.status_code})"
                continue

            articles = []
            total = 0
            for entry in feed.entries:
                total += 1
                pub_date = parse_date(entry)

                if pub_date and pub_date < ONE_WEEK_AGO:
                    continue

                content = ""
                for field in ("content", "summary", "description"):
                    val = getattr(entry, field, None)
                    if val:
                        if isinstance(val, list):
                            val = val[0].get("value", "") if val else ""
                        content = clean_html(val)
                        if content:
                            break

                title = clean_html(entry.get("title", ""))
                link = entry.get("link", "")

                if not title:
                    continue

                articles.append(
                    {
                        "source": source["name"],
                        "title": title,
                        "url": link,
                        "content": content[:1200],
                        "pub_date": pub_date.strftime("%Y-%m-%d") if pub_date else "unknown",
                        "lang": source.get("lang", "es"),
                    }
                )

            detail = f"OK - {len(articles)}/{total} articles in past week (from {url})"
            return articles, None, detail, len(articles)

        except requests.exceptions.HTTPError as e:
            last_error = f"HTTP {e.response.status_code}: {url}"
        except requests.exceptions.ConnectionError:
            last_error = f"Connection error: {url}"
        except requests.exceptions.Timeout:
            last_error = f"Timeout: {url}"
        except Exception as e:
            last_error = f"Error ({type(e).__name__}): {e}"

    return [], last_error, None, 0


# ─────────────────────────────────────────
# Claude Integration
# ─────────────────────────────────────────

RELEVANCE_SYSTEM = """You are a senior analyst filtering news for a Mexico financial & fintech weekly briefing.

RELEVANT topics (any of these qualifies):
1. Mexico macroeconomy: GDP, inflation, Banxico interest rates, peso (MXN/USD), employment, government budget, trade balance
2. Mexico politics & policy: legislation, regulations, CNBV/Banxico announcements, tax changes, anything impacting finance or business
3. Mexico fintech players (products, funding, partnerships, campaigns, layoffs, expansions):
   Nu, Mercado Pago, Stori, Plata, Kueski, Klar, Openbank, Uala (Ualá), Aplazo, Revolut, Clip, Conekta, Bitso, Kushki
4. Mexico traditional banks (same dimensions):
   BBVA Mexico, Santander Mexico, Bancoppel, Banco Azteca, Citibanamex, Scotiabank Mexico, HSBC Mexico, Inbursa
5. Financial products in Mexico: cash loans (crédito personal), credit cards, BNPL/meses sin intereses, Cuenta/savings accounts, neobanks, payments, remittances
6. International events with CLEAR, DIRECT impact on Mexico:
   US-Mexico tariffs/trade, USD/MXN, oil prices (Mexico is an oil exporter), nearshoring, remittances from the US
7. Overseas/international activities of Mexico's core fintech or banking players — even if the event happens outside Mexico:
   e.g. Nu expanding to the US, Plata launching in another country, Mercado Pago entering a new market, Clip or Bitso raising a global round.
   Core players: Nu, Mercado Pago, Stori, Plata, Kueski, Klar, Openbank, Uala, Aplazo, Revolut, Clip, Conekta, Bitso, Kushki,
   BBVA Mexico, Santander Mexico, Bancoppel, Banco Azteca, Citibanamex, Scotiabank Mexico, HSBC Mexico, Inbursa

NOT relevant: general world politics without Mexico link, sports, entertainment, purely US/EU domestic news, tech news unrelated to finance."""

GROUPING_SYSTEM = """You are creating a concise Mexico financial & fintech weekly newsletter.
Write in English. Be factual, precise, and concise. Focus on what matters to a financial professional.

CORE PLAYERS (well-known to readers — NO intro needed):
- Fintech: Nu (Nubank), Mercado Pago, Stori, Plata, Kueski, Klar, Openbank, Uala (Ualá), Aplazo, Revolut, Clip, Conekta, Bitso, Kushki
- Banking: BBVA Mexico, Santander Mexico, Bancoppel, Banco Azteca, Citibanamex, Scotiabank Mexico, HSBC Mexico, Inbursa

For any OTHER financial company first mentioned in a summary, immediately after its name add a short parenthetical identifying its main business, e.g.:
  "Covalto (SME digital lender focused on working-capital loans) announced..."
  "Fintual (robo-advisor platform for retail investors in Mexico and Chile) reported..."
Keep the parenthetical to one concise phrase — do not repeat it if the company is mentioned again."""


def filter_relevant(articles, client, batch_size=25):
    """Use Claude to filter articles relevant to Mexico finance/fintech topics."""
    relevant = []
    total_batches = (len(articles) + batch_size - 1) // batch_size

    for i in range(0, len(articles), batch_size):
        batch = articles[i : i + batch_size]
        batch_num = i // batch_size + 1
        print(f"    Relevance check: batch {batch_num}/{total_batches} ({len(batch)} articles)...")

        items_text = "\n".join(
            f"[{j}] TITLE: {a['title']}\n    SOURCE: {a['source']} | DATE: {a['pub_date']}\n    SNIPPET: {a['content'][:250]}"
            for j, a in enumerate(batch)
        )

        prompt = (
            f"Review these {len(batch)} news articles. Return ONLY the IDs (0-based integers) "
            "of articles that are RELEVANT per your system instructions.\n\n"
            f"{items_text}\n\n"
            "Return a JSON array of integers only. Example: [0, 2, 5]\n"
            "If none are relevant, return: []"
        )

        try:
            resp = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=400,
                system=RELEVANCE_SYSTEM,
                messages=[{"role": "user", "content": prompt}],
            )
            text = resp.content[0].text.strip()
            match = re.search(r"\[[\d,\s]*\]", text)
            if match:
                ids = json.loads(match.group())
                for idx in ids:
                    if 0 <= idx < len(batch):
                        relevant.append(batch[idx])
        except Exception as e:
            print(f"    [WARN] Relevance check failed for batch {batch_num}: {e}")
            relevant.extend(batch)

        time.sleep(0.5)

    return relevant


def group_and_summarize(articles, client):
    """Use Claude to group duplicate stories and write English summaries."""
    if not articles:
        return []

    items = []
    for i, a in enumerate(articles):
        items.append(
            {
                "id": i,
                "source": a["source"],
                "title": a["title"],
                "url": a["url"],
                "date": a["pub_date"],
                "snippet": a["content"][:400],
            }
        )

    chunk_size = 40
    all_groups = []

    for chunk_start in range(0, len(items), chunk_size):
        chunk = items[chunk_start : chunk_start + chunk_size]
        chunk_label = f"{chunk_start+1}-{chunk_start+len(chunk)}"
        print(f"  Grouping & summarizing articles {chunk_label}/{len(items)}...")

        prompt = (
            f"Here are {len(chunk)} relevant news articles about Mexico finance/fintech from the past week.\n\n"
            f"ARTICLES (JSON):\n{json.dumps(chunk, ensure_ascii=False, indent=2)}\n\n"
            "Your tasks:\n"
            "1. GROUP articles that cover the SAME event/story (even if from different sources)\n"
            "2. For each group, write a 100-150 word English summary covering: what happened, key numbers/facts, and significance\n"
            "3. Assign ONE category per group:\n"
            '   - "macro": Macroeconomy, inflation, interest rates, peso, government policy, regulations\n'
            '   - "financial": Any fintech or traditional bank news (Nu, Mercado Pago, BBVA, Santander, etc.)\n'
            '   - "international": International events with specific impact on Mexico\n\n'
            "Return ONLY a valid JSON array. Each element:\n"
            '{"title":"Clear English headline (max 12 words)","summary":"100-150 word English summary",'
            '"category":"macro|financial|international","article_ids":[list of IDs],'
            '"sources":[{"name":"Source Name","url":"article URL"}]}\n\n'
            "Important: article_ids uses the \"id\" field from the input JSON above."
        )

        try:
            resp = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=6000,
                system=GROUPING_SYSTEM,
                messages=[{"role": "user", "content": prompt}],
            )
            text = resp.content[0].text.strip()
            json_match = re.search(r"\[.*\]", text, re.DOTALL)
            if json_match:
                groups = json.loads(json_match.group())
                all_groups.extend(groups)
            else:
                print(f"  [WARN] Could not parse grouping response for chunk {chunk_label}")
        except json.JSONDecodeError as e:
            print(f"  [WARN] JSON parse error for chunk {chunk_label}: {e}")
        except Exception as e:
            print(f"  [WARN] Grouping failed for chunk {chunk_label}: {e}")

        time.sleep(1)

    return all_groups


# ─────────────────────────────────────────
# HTML Generation v2
# ─────────────────────────────────────────

CATEGORY_META = {
    "macro": {
        "label": "Macro & Policy",
        "icon": "&#127963;",   # 🏛
        "accent": "#2563eb",
        "text": "#1e3a6e",
    },
    "financial": {
        "label": "Financial Players",
        "icon": "&#128179;",   # 💳
        "accent": "#0891b2",   # teal
        "text": "#164e63",
    },
    "international": {
        "label": "International Impact",
        "icon": "&#127760;",   # 🌐
        "accent": "#9333ea",
        "text": "#4c1d95",
    },
}

CAT_ORDER = ["macro", "financial", "international"]

# Remap legacy category values from old runs / old Claude responses
CAT_REMAP = {"fintech": "financial", "banking": "financial"}


def _esc(s):
    """HTML-escape for safe insertion into text and attribute values."""
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _build_card(item, story_id, category):
    meta = CATEGORY_META.get(category, CATEGORY_META["macro"])
    accent = meta["accent"]

    sources_html = ""
    for src in item.get("sources", []):
        url = _esc(src.get("url") or "#")
        name = _esc(src.get("name") or "Source")
        sources_html += (
            f'<a href="{url}" target="_blank" rel="noopener" class="source-link">{name}</a>'
        )

    title   = _esc(item.get("title") or "Untitled")
    summary = _esc(item.get("summary") or "")

    return (
        f'<div class="card" id="{story_id}" data-category="{category}" style="--card-accent:{accent}">'
        f'<div class="card-toolbar">'
        f'<span class="drag-handle" title="Drag to reorder">&#8942;&#8942;</span>'
        f'<div class="card-btns">'
        f'<button class="btn-icon btn-edit" onclick="openEdit(\'{story_id}\')" title="Edit">&#9998;</button>'
        f'<button class="btn-icon btn-delete" onclick="deleteCard(\'{story_id}\')" title="Delete">&#215;</button>'
        f'</div></div>'
        f'<div class="card-header-row" onclick="toggleCard(event)">'
        f'<span class="card-toggle-icon">&#9658;</span>'
        f'<h3 class="card-title">{title}</h3>'
        f'</div>'
        f'<p class="card-summary">{summary}</p>'
        f'<div class="card-sources"><span class="sources-label">Sources:</span>{sources_html}</div>'
        f'</div>\n'
    )


def _build_sections(grouped):
    by_cat = defaultdict(list)
    for item in grouped:
        cat = item.get("category", "macro")
        cat = CAT_REMAP.get(cat, cat)   # fintech/banking → financial
        if cat not in CATEGORY_META:
            cat = "macro"
        by_cat[cat].append(item)

    html = ""
    counter = [0]
    for cat in CAT_ORDER:
        items = by_cat.get(cat, [])
        meta  = CATEGORY_META[cat]
        accent = meta["accent"]

        cards_html = ""
        for item in items:
            sid = f"story-{counter[0]}"
            counter[0] += 1
            cards_html += _build_card(item, sid, cat)

        html += (
            f'<section class="cat-section" id="section-{cat}" data-category="{cat}" style="--cat-accent:{accent}">\n'
            f'<div class="cat-header">'
            f'<span class="cat-icon">{meta["icon"]}</span>'
            f'<span class="cat-title">{meta["label"]}</span>'
            f'<span class="cat-badge" id="badge-{cat}">{len(items)}</span>'
            f'</div>\n'
            f'<div class="cards-list" id="cards-{cat}">{cards_html}</div>\n'
            f'<button class="btn-add" onclick="openAdd(\'{cat}\')">+ Add Story</button>\n'
            f'</section>\n'
        )

    return html


def get_html_template():
    """Return the complete HTML/CSS/JS template with PLACEHOLDER_ tokens."""
    return """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Mexico Weekly Newsletter &mdash; PLACEHOLDER_WEEK_END</title>
<script src="https://cdn.jsdelivr.net/npm/sortablejs@1.15.2/Sortable.min.js"></script>
<style>
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

:root {
  --body-bg: #eef1f6;
  --card-bg: #ffffff;
  --text: #111827;
  --text-muted: #6b7280;
  --border: #e5e7eb;
  --radius: 10px;
  --shadow: 0 1px 3px rgba(0,0,0,0.07), 0 1px 2px rgba(0,0,0,0.04);
  --shadow-hover: 0 6px 18px rgba(0,0,0,0.11);
  --font: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
  --mono: "JetBrains Mono", "Fira Code", "Courier New", monospace;
}

body {
  font-family: var(--font);
  background: var(--body-bg);
  color: var(--text);
  line-height: 1.6;
  min-height: 100vh;
}

/* ══════════════════════════════════════════
   HEADER  —  Dark Fintech + Mexico Flag
══════════════════════════════════════════ */
.header {
  position: relative;
  background: linear-gradient(158deg, #070e1c 0%, #0d1f3c 50%, #0a1628 100%);
  color: #fff;
  overflow: hidden;
}

/* Mexico flag strip */
.flag-strip {
  height: 9px;
  background: linear-gradient(90deg, #006847 33.33%, #ffffff 33.33%, #ffffff 66.66%, #ce1126 66.66%);
  position: relative;
  z-index: 2;
}

/* Tech dot-grid pattern */
.header-inner {
  position: relative;
  padding: 28px 40px 36px;
  background-image: radial-gradient(circle, rgba(255,255,255,0.055) 1px, transparent 1px);
  background-size: 26px 26px;
}

/* Small stats — top left */
.header-stats {
  position: relative;
  z-index: 1;
  display: flex;
  flex-wrap: wrap;
  gap: 6px;
  margin-bottom: 30px;
}
.stat-chip {
  font-family: var(--mono);
  font-size: 0.68rem;
  background: rgba(255,255,255,0.07);
  border: 1px solid rgba(255,255,255,0.13);
  color: rgba(255,255,255,0.65);
  padding: 3px 10px;
  border-radius: 4px;
  letter-spacing: 0.35px;
}
.stat-chip strong {
  color: rgba(255,255,255,0.92);
  font-weight: 700;
}
.stat-sep {
  color: rgba(255,255,255,0.2);
  font-size: 0.7rem;
  align-self: center;
  font-family: var(--mono);
}

/* Title */
.header-center {
  position: relative;
  z-index: 1;
  text-align: center;
  margin-bottom: 28px;
  padding-top: 6px;
}
.header-title {
  font-size: 2.2rem;
  font-weight: 800;
  letter-spacing: -0.5px;
  background: linear-gradient(135deg, #ffffff 0%, rgba(255,255,255,0.80) 100%);
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  background-clip: text;
  margin-bottom: 9px;
}
.header-sub {
  font-size: 0.85rem;
  color: rgba(255,255,255,0.48);
  letter-spacing: 0.6px;
  text-transform: uppercase;
}

/* Category nav pills */
.cat-nav {
  display: flex;
  gap: 8px;
  justify-content: center;
  flex-wrap: wrap;
  position: relative;
  z-index: 1;
}
.cat-nav-pill {
  display: inline-flex;
  align-items: center;
  gap: 5px;
  padding: 6px 14px;
  border-radius: 99px;
  font-size: 0.77rem;
  font-weight: 500;
  text-decoration: none;
  background: rgba(255,255,255,0.07);
  color: rgba(255,255,255,0.7);
  border: 1px solid rgba(255,255,255,0.11);
  transition: background 0.18s, color 0.18s, border-color 0.18s;
  white-space: nowrap;
}
.cat-nav-pill:hover {
  background: rgba(255,255,255,0.14);
  color: #fff;
  border-color: rgba(255,255,255,0.22);
}

/* ══════════════════════════════════════════
   MAIN CONTENT
══════════════════════════════════════════ */
.container {
  max-width: 860px;
  margin: 0 auto;
  padding: 40px 24px 80px;
}

/* Category sections */
.cat-section {
  margin-bottom: 52px;
}
.cat-header {
  display: flex;
  align-items: center;
  gap: 10px;
  margin-bottom: 16px;
  padding-bottom: 12px;
  border-bottom: 2px solid var(--cat-accent);
}
.cat-icon { font-size: 1.25rem; }
.cat-title {
  font-size: 1.15rem;
  font-weight: 700;
  color: #1a1a2e;
  flex: 1;
}
.cat-badge {
  font-family: var(--mono);
  font-size: 0.7rem;
  font-weight: 700;
  background: var(--cat-accent);
  color: #fff;
  padding: 2px 9px;
  border-radius: 99px;
  min-width: 26px;
  text-align: center;
}

/* Cards container */
.cards-list {
  display: flex;
  flex-direction: column;
  gap: 10px;
  min-height: 44px;
  padding: 2px 0;
}

/* Individual card */
.card {
  background: var(--card-bg);
  border-radius: var(--radius);
  border-left: 4px solid var(--card-accent, #2563eb);
  padding: 12px 16px 14px;
  box-shadow: var(--shadow);
  position: relative;
  transition: box-shadow 0.18s, transform 0.15s;
}
.card:hover {
  box-shadow: var(--shadow-hover);
  transform: translateY(-1px);
}

/* Card toolbar (edit/delete/drag) — visible on hover */
.card-toolbar {
  display: flex;
  align-items: center;
  justify-content: space-between;
  margin-bottom: 9px;
  height: 22px;
  opacity: 0;
  pointer-events: none;
  transition: opacity 0.15s;
}
body.edit-mode .card:hover .card-toolbar { opacity: 1; pointer-events: auto; }

.drag-handle {
  color: #d1d5db;
  font-size: 0.95rem;
  cursor: grab;
  padding: 2px 4px;
  border-radius: 4px;
  letter-spacing: -3px;
  user-select: none;
  transition: color 0.15s;
}
.drag-handle:active { cursor: grabbing; }
.drag-handle:hover  { color: #6b7280; }

.card-btns { display: flex; gap: 4px; }

.btn-icon {
  width: 26px;
  height: 26px;
  border: none;
  border-radius: 6px;
  cursor: pointer;
  font-size: 0.88rem;
  display: flex;
  align-items: center;
  justify-content: center;
  background: transparent;
  color: #9ca3af;
  transition: background 0.15s, color 0.15s;
}
.btn-edit:hover   { background: #eff6ff; color: #2563eb; }
.btn-delete:hover { background: #fef2f2; color: #dc2626; }

.card-title {
  font-size: 0.98rem;
  font-weight: 700;
  color: #111827;
  margin-bottom: 8px;
  line-height: 1.35;
}
.card-summary {
  font-size: 0.875rem;
  color: #374151;
  line-height: 1.72;
  margin-bottom: 13px;
}
.card-sources {
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  gap: 6px;
  padding-top: 10px;
  border-top: 1px solid #f3f4f6;
}
.sources-label {
  font-size: 0.68rem;
  font-weight: 700;
  color: #9ca3af;
  text-transform: uppercase;
  letter-spacing: 0.6px;
}
.source-link {
  font-size: 0.7rem;
  color: #2563eb;
  text-decoration: none;
  background: #eff6ff;
  padding: 2px 9px;
  border-radius: 99px;
  border: 1px solid #bfdbfe;
  white-space: nowrap;
  transition: background 0.15s;
}
.source-link:hover { background: #dbeafe; text-decoration: underline; }

/* Expand/Collapse All button */
.btn-toggle-all {
  position: fixed;
  bottom: 28px;
  right: 28px;
  z-index: 200;
  display: inline-flex;
  align-items: center;
  gap: 6px;
  padding: 8px 18px;
  border-radius: 99px;
  font-size: 0.78rem;
  font-weight: 600;
  cursor: pointer;
  border: none;
  background: #1e293b;
  color: #f1f5f9;
  box-shadow: 0 4px 14px rgba(0,0,0,0.22);
  transition: background 0.15s, box-shadow 0.15s, transform 0.1s;
  white-space: nowrap;
}
.btn-toggle-all:hover {
  background: #0f172a;
  box-shadow: 0 6px 20px rgba(0,0,0,0.30);
  transform: translateY(-1px);
}

/* ── Collapse / Expand ── */
.card-header-row {
  display: flex;
  align-items: flex-start;
  gap: 7px;
  cursor: pointer;
  user-select: none;
  margin-bottom: 0;
}
.card-toggle-icon {
  color: #c4c9d4;
  font-size: 0.58rem;
  flex-shrink: 0;
  margin-top: 4px;
  transition: transform 0.2s, color 0.2s;
  display: inline-block;
}
.card.expanded .card-toggle-icon {
  transform: rotate(90deg);
  color: var(--card-accent);
}
.card-summary,
.card-sources {
  display: none;
}
.card.expanded .card-summary { display: block; }
.card.expanded .card-sources { display: flex; }
.card.expanded .card-header-row { margin-bottom: 8px; }

/* Sortable drag states */
.card-ghost  { opacity: 0.3 !important; background: #e0e7ff; border-radius: var(--radius); }
.card-chosen { box-shadow: 0 14px 36px rgba(0,0,0,0.22) !important; transform: rotate(0.6deg) scale(1.01) !important; }
.card-drag   { box-shadow: 0 18px 44px rgba(0,0,0,0.28) !important; }

/* Add Story button */
.btn-add {
  display: none;
  align-items: center;
  justify-content: center;
  gap: 6px;
  margin-top: 10px;
  padding: 8px 16px;
  width: 100%;
  border: 2px dashed var(--cat-accent);
  border-radius: 8px;
  background: transparent;
  color: var(--cat-accent);
  font-size: 0.8rem;
  font-weight: 600;
  cursor: pointer;
  opacity: 0.6;
  transition: opacity 0.18s, background 0.18s, color 0.18s;
}
.btn-add:hover {
  opacity: 1;
  background: var(--cat-accent);
  color: #fff;
}

/* ══════════════════════════════════════════
   MODAL
══════════════════════════════════════════ */
.modal-overlay {
  position: fixed;
  inset: 0;
  background: rgba(0,0,0,0.6);
  z-index: 200;
  display: none;
  backdrop-filter: blur(3px);
}
.modal-overlay.active { display: block; }

.modal {
  position: fixed;
  top: 50%;
  left: 50%;
  transform: translate(-50%, -50%) scale(0.95);
  background: #fff;
  border-radius: 14px;
  width: min(580px, 94vw);
  max-height: 88vh;
  overflow-y: auto;
  z-index: 201;
  box-shadow: 0 24px 64px rgba(0,0,0,0.35);
  display: none;
  flex-direction: column;
  transition: transform 0.2s;
}
.modal.active {
  display: flex;
  transform: translate(-50%, -50%) scale(1);
}

.modal-head {
  padding: 18px 20px 14px;
  border-bottom: 1px solid #e5e7eb;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 12px;
  flex-shrink: 0;
}
.modal-heading {
  font-size: 1rem;
  font-weight: 700;
  color: #111827;
}
.modal-close {
  width: 28px; height: 28px;
  border: none; border-radius: 6px;
  background: #f3f4f6; color: #6b7280;
  cursor: pointer; font-size: 1rem;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
  transition: background 0.15s;
}
.modal-close:hover { background: #e5e7eb; color: #374151; }

.modal-body { padding: 18px 20px; flex: 1; }

.form-group { margin-bottom: 15px; }
.form-label {
  display: block;
  font-size: 0.73rem;
  font-weight: 700;
  color: #4b5563;
  text-transform: uppercase;
  letter-spacing: 0.5px;
  margin-bottom: 5px;
}
.input-field {
  width: 100%;
  padding: 8px 11px;
  border: 1px solid #d1d5db;
  border-radius: 7px;
  font-size: 0.875rem;
  color: #111827;
  background: #fafafa;
  outline: none;
  font-family: inherit;
  transition: border-color 0.18s, box-shadow 0.18s;
}
.input-field:focus {
  border-color: #2563eb;
  box-shadow: 0 0 0 3px rgba(37,99,235,0.12);
  background: #fff;
}
textarea.input-field {
  resize: vertical;
  min-height: 108px;
  line-height: 1.65;
}

.sources-header {
  font-size: 0.73rem;
  font-weight: 700;
  color: #4b5563;
  text-transform: uppercase;
  letter-spacing: 0.5px;
  margin-bottom: 7px;
}
#modal-sources { display: flex; flex-direction: column; gap: 6px; margin-bottom: 8px; }
.src-row {
  display: grid;
  grid-template-columns: 1fr 2fr 28px;
  gap: 6px;
  align-items: center;
}
.src-row .input-field { padding: 6px 10px; font-size: 0.82rem; }
.btn-rm-src {
  width: 28px; height: 28px;
  border: 1px solid #fca5a5; background: #fef2f2; color: #dc2626;
  border-radius: 6px; cursor: pointer; font-size: 0.85rem;
  display: flex; align-items: center; justify-content: center;
  transition: background 0.15s;
}
.btn-rm-src:hover { background: #fee2e2; }

.btn-add-src {
  display: inline-flex; align-items: center; gap: 5px;
  padding: 5px 12px;
  border: 1px dashed #9ca3af; border-radius: 6px;
  background: transparent; color: #6b7280;
  font-size: 0.78rem; cursor: pointer;
  transition: border-color 0.15s, color 0.15s;
}
.btn-add-src:hover { border-color: #4b5563; color: #374151; }

.modal-foot {
  padding: 13px 20px;
  border-top: 1px solid #e5e7eb;
  display: flex;
  justify-content: flex-end;
  gap: 8px;
  flex-shrink: 0;
}
.btn-cancel {
  padding: 8px 18px;
  border: 1px solid #d1d5db; border-radius: 7px;
  background: #fff; color: #374151;
  font-size: 0.875rem; font-weight: 500; cursor: pointer;
  transition: background 0.15s;
}
.btn-cancel:hover { background: #f9fafb; }
.btn-save {
  padding: 8px 22px;
  border: none; border-radius: 7px;
  background: #2563eb; color: #fff;
  font-size: 0.875rem; font-weight: 600; cursor: pointer;
  transition: background 0.18s;
}
.btn-save:hover { background: #1d4ed8; }

/* ══════════════════════════════════════════
   FOOTER
══════════════════════════════════════════ */
.footer {
  text-align: center;
  font-size: 0.73rem;
  color: #9ca3af;
  padding: 22px 24px;
  border-top: 1px solid #e5e7eb;
  background: #fff;
}

/* ══════════════════════════════════════════
   EDITOR MODE
══════════════════════════════════════════ */
body.edit-mode .btn-add { display: flex; }

.btn-editor-lock {
  position: fixed; bottom: 28px; left: 28px; z-index: 200;
  width: 36px; height: 36px; border-radius: 99px;
  border: 1px solid rgba(30,41,59,0.15);
  background: rgba(255,255,255,0.75); color: #94a3b8;
  font-size: 0.9rem; cursor: pointer;
  display: flex; align-items: center; justify-content: center;
  backdrop-filter: blur(4px);
  transition: background 0.15s, color 0.15s, border-color 0.15s;
}
.btn-editor-lock:hover { background: #fff; color: #475569; border-color: #94a3b8; }
body.edit-mode .btn-editor-lock { display: none; }

.edit-mode-bar {
  display: none; position: fixed; bottom: 20px; left: 20px; z-index: 200;
  align-items: center; gap: 6px;
  background: #1e293b; padding: 6px 8px 6px 14px; border-radius: 99px;
  box-shadow: 0 4px 14px rgba(0,0,0,0.22);
}
body.edit-mode .edit-mode-bar { display: flex; }
.edit-mode-label {
  font-size: 0.68rem; font-weight: 700; color: #64748b;
  text-transform: uppercase; letter-spacing: 0.5px; padding-right: 4px;
}
.btn-publish {
  background: #16a34a; color: #fff; border: none;
  padding: 5px 14px; border-radius: 99px;
  font-size: 0.75rem; font-weight: 700; cursor: pointer;
  transition: background 0.15s;
}
.btn-publish:hover { background: #15803d; }
.btn-publish:disabled { background: #4ade80; cursor: default; }
.btn-exit-edit {
  background: transparent; color: #64748b;
  border: 1px solid rgba(255,255,255,0.12);
  padding: 5px 12px; border-radius: 99px;
  font-size: 0.75rem; font-weight: 600; cursor: pointer;
  transition: color 0.15s, border-color 0.15s;
}
.btn-exit-edit:hover { color: #e2e8f0; border-color: rgba(255,255,255,0.3); }

.pw-overlay {
  position: fixed; inset: 0; background: rgba(0,0,0,0.55);
  z-index: 400; display: none; backdrop-filter: blur(3px);
}
.pw-overlay.active { display: block; }
.pw-modal {
  position: fixed; top: 50%; left: 50%;
  transform: translate(-50%,-50%) scale(0.95);
  background: #fff; border-radius: 14px;
  width: min(360px, calc(100vw - 32px));
  z-index: 401; box-shadow: 0 20px 60px rgba(0,0,0,0.18);
  opacity: 0; pointer-events: none;
  transition: transform 0.18s, opacity 0.18s;
}
.pw-modal.active { transform: translate(-50%,-50%) scale(1); opacity: 1; pointer-events: auto; }
.pw-modal-head {
  display: flex; align-items: center; justify-content: space-between;
  padding: 18px 20px 12px; font-weight: 700; font-size: 0.95rem; color: #111827;
}
.pw-modal-body { padding: 0 20px 20px; }
.pw-error { color: #dc2626; font-size: 0.78rem; margin-top: 8px; display: none; }

/* ══════════════════════════════════════════
   RESPONSIVE
══════════════════════════════════════════ */
@media (max-width: 600px) {
  .header-inner { padding: 22px 20px 28px; }
  .header-title { font-size: 1.5rem; }
  .container { padding: 28px 14px 60px; }
  .src-row { grid-template-columns: 1fr 28px; }
  .src-row .src-url { display: none; }
}
</style>
</head>
<body>

<!-- ══ HEADER ══ -->
<header class="header">
  <div class="flag-strip"></div>
  <div class="header-inner">
    <!-- Stats — small, top-left -->
    <div class="header-stats">
      <span class="stat-chip"><strong>PLACEHOLDER_TOTAL_STORIES</strong> stories</span>
      <span class="stat-sep">/</span>
      <span class="stat-chip"><strong>PLACEHOLDER_TOTAL_RELEVANT</strong> relevant</span>
      <span class="stat-sep">/</span>
      <span class="stat-chip"><strong>PLACEHOLDER_TOTAL_RAW</strong> scanned</span>
      <span class="stat-sep">/</span>
      <span class="stat-chip"><strong>PLACEHOLDER_ACTIVE_SOURCES</strong> sources</span>
    </div>

    <!-- Title -->
    <div class="header-center">
      <h1 class="header-title">Mexico Weekly Newsletter</h1>
      <p class="header-sub">PLACEHOLDER_WEEK_START &ndash; PLACEHOLDER_WEEK_END &nbsp;&bull;&nbsp; Macro &bull; Finance &bull; Fintech</p>
    </div>

    <!-- Category nav -->
    <nav class="cat-nav">PLACEHOLDER_CAT_NAV</nav>
  </div>
</header>

<!-- ══ STORIES ══ -->
<main class="container">
PLACEHOLDER_SECTIONS
</main>

<!-- ══ FOOTER ══ -->
<footer class="footer">
  Generated by Mexico News Aggregator &nbsp;&bull;&nbsp; Powered by Claude (Anthropic) &nbsp;&bull;&nbsp; PLACEHOLDER_GENERATED_AT
</footer>

<!-- ══ MODAL ══ -->
<div class="modal-overlay" id="modal-overlay" onclick="closeModal()"></div>
<div class="modal" id="modal" role="dialog" aria-modal="true" aria-labelledby="modal-heading">
  <div class="modal-head">
    <span class="modal-heading" id="modal-heading">Edit Story</span>
    <button class="modal-close" onclick="closeModal()" title="Close">&#215;</button>
  </div>
  <div class="modal-body">
    <div class="form-group">
      <label class="form-label" for="modal-title-input">Headline</label>
      <input id="modal-title-input" type="text" class="input-field" placeholder="Clear, concise headline (max 12 words)"/>
    </div>
    <div class="form-group">
      <label class="form-label" for="modal-summary">Summary (100&ndash;150 words)</label>
      <textarea id="modal-summary" class="input-field" placeholder="Write the 100-150 word summary here..."></textarea>
    </div>
    <div class="form-group">
      <div class="sources-header">Sources</div>
      <div id="modal-sources"></div>
      <button type="button" class="btn-add-src" onclick="addSourceRow('','')">&#43; Add source</button>
    </div>
  </div>
  <div class="modal-foot">
    <button class="btn-cancel" onclick="closeModal()">Cancel</button>
    <button class="btn-save" onclick="saveCard()">Save</button>
  </div>
</div>

<!-- ══ JAVASCRIPT ══ -->
<script>
/* ── State ── */
var editCardId   = null;
var addCategory  = null;

/* ── Category metadata (must match Python) ── */
var ACCENTS = {
  macro:         '#2563eb',
  financial:     '#0891b2',
  international: '#9333ea'
};
var CAT_LABELS = {
  macro:         'Macro & Policy',
  financial:     'Financial Players',
  international: 'International Impact'
};
var CAT_ICONS = {
  macro:         '&#127963;',
  financial:     '&#128179;',
  international: '&#127760;'
};

/* ── Helpers ── */
function escHtml(s) {
  return (s || '')
    .replace(/&/g,  '&amp;')
    .replace(/</g,  '&lt;')
    .replace(/>/g,  '&gt;')
    .replace(/"/g,  '&quot;');
}

/* ── Toggle all cards ── */
var allExpanded = false;
function toggleAll() {
  allExpanded = !allExpanded;
  document.querySelectorAll('.card').forEach(function(card) {
    if (allExpanded) card.classList.add('expanded');
    else             card.classList.remove('expanded');
  });
  var btn = document.getElementById('btn-toggle-all');
  if (btn) btn.innerHTML = allExpanded ? '&#9650; Collapse All' : '&#9660; Expand All';
}

/* ── Collapse / Expand ── */
function toggleCard(event) {
  var card = event.currentTarget.closest('.card');
  if (card) card.classList.toggle('expanded');
}

/* ── Modal open: edit existing card ── */
function openEdit(cardId) {
  editCardId  = cardId;
  addCategory = null;
  var card = document.getElementById(cardId);
  if (!card) return;

  document.getElementById('modal-heading').textContent = 'Edit Story';
  document.getElementById('modal-title-input').value =
    card.querySelector('.card-title').textContent;
  document.getElementById('modal-summary').value =
    card.querySelector('.card-summary').textContent;

  var srcsEl = document.getElementById('modal-sources');
  srcsEl.innerHTML = '';
  var links = card.querySelectorAll('.source-link');
  if (links.length > 0) {
    links.forEach(function(a) { addSourceRow(a.textContent.trim(), a.href); });
  } else {
    addSourceRow('', '');
  }
  showModal();
}

/* ── Modal open: add new card in a category ── */
function openAdd(category) {
  editCardId  = null;
  addCategory = category;
  document.getElementById('modal-heading').innerHTML =
    'Add Story &mdash; ' + (CAT_ICONS[category] || '') + ' ' + (CAT_LABELS[category] || category);
  document.getElementById('modal-title-input').value   = '';
  document.getElementById('modal-summary').value       = '';
  var srcsEl = document.getElementById('modal-sources');
  srcsEl.innerHTML = '';
  addSourceRow('', '');
  showModal();
}

function showModal() {
  document.getElementById('modal-overlay').classList.add('active');
  document.getElementById('modal').classList.add('active');
  setTimeout(function() {
    var t = document.getElementById('modal-title-input');
    if (t) t.focus();
  }, 60);
}

function closeModal() {
  document.getElementById('modal-overlay').classList.remove('active');
  document.getElementById('modal').classList.remove('active');
  editCardId  = null;
  addCategory = null;
}

/* ── Add a source input row inside the modal ── */
function addSourceRow(name, url) {
  var div = document.createElement('div');
  div.className = 'src-row';
  div.innerHTML =
    '<input type="text" class="src-name input-field" placeholder="Source name" value="' + escHtml(name || '') + '">' +
    '<input type="url"  class="src-url  input-field" placeholder="https://..."  value="' + escHtml(url  || '') + '">' +
    '<button type="button" class="btn-rm-src" onclick="this.closest(&#39;.src-row&#39;).remove()" title="Remove">&#215;</button>';
  document.getElementById('modal-sources').appendChild(div);
}

/* ── Save card (edit or create) ── */
function saveCard() {
  var title   = document.getElementById('modal-title-input').value.trim();
  var summary = document.getElementById('modal-summary').value.trim();
  if (!title)   { alert('Please enter a headline.'); return; }
  if (!summary) { alert('Please enter a summary.');  return; }

  /* Build sources HTML */
  var srcHtml = '';
  document.querySelectorAll('.src-row').forEach(function(row) {
    var n = row.querySelector('.src-name').value.trim();
    var u = row.querySelector('.src-url').value.trim();
    if (n) {
      srcHtml += '<a href="' + (u || '#') + '" target="_blank" rel="noopener" class="source-link">' + escHtml(n) + '</a>';
    }
  });

  if (editCardId) {
    /* ── Update existing card ── */
    var card = document.getElementById(editCardId);
    card.querySelector('.card-title').textContent   = title;
    card.querySelector('.card-summary').textContent = summary;
    card.querySelector('.card-sources').innerHTML   =
      '<span class="sources-label">Sources:</span>' + srcHtml;

  } else {
    /* ── Create new card ── */
    var cat    = addCategory;
    var accent = ACCENTS[cat] || '#2563eb';
    var id     = 'story-' + Date.now();

    var card = document.createElement('div');
    card.className        = 'card';
    card.id               = id;
    card.dataset.category = cat;
    card.style.cssText    = '--card-accent:' + accent;

    card.innerHTML =
      '<div class="card-toolbar">' +
        '<span class="drag-handle" title="Drag to reorder">&#8942;&#8942;</span>' +
        '<div class="card-btns">' +
          '<button class="btn-icon btn-edit"   onclick="openEdit(&#39;' + id + '&#39;)"   title="Edit">&#9998;</button>' +
          '<button class="btn-icon btn-delete" onclick="deleteCard(&#39;' + id + '&#39;)" title="Delete">&#215;</button>' +
        '</div>' +
      '</div>' +
      '<div class="card-header-row" onclick="toggleCard(event)">' +
        '<span class="card-toggle-icon">&#9658;</span>' +
        '<h3 class="card-title">' + escHtml(title) + '</h3>' +
      '</div>' +
      '<p  class="card-summary">' + escHtml(summary) + '</p>' +
      '<div class="card-sources"><span class="sources-label">Sources:</span>' + srcHtml + '</div>';

    card.classList.add('expanded');   // new cards start expanded
    var container = document.getElementById('cards-' + cat);
    if (container) container.appendChild(card);
    updateCounts();
  }

  closeModal();
}

/* ── Delete a card ── */
function deleteCard(cardId) {
  if (!confirm('Delete this story?')) return;
  var card = document.getElementById(cardId);
  if (card) {
    card.remove();
    updateCounts();
  }
}

/* ── Refresh the badge counts in each category header ── */
function updateCounts() {
  ['macro', 'financial', 'international'].forEach(function(cat) {
    var container = document.getElementById('cards-' + cat);
    var badge     = document.getElementById('badge-' + cat);
    if (container && badge) {
      badge.textContent = container.querySelectorAll('.card').length;
    }
  });
}

/* ── Init Sortable.js drag-and-drop ── */
function initSortable() {
  if (typeof Sortable === 'undefined') {
    console.warn('Sortable.js not loaded — drag-and-drop disabled.');
    return;
  }
  ['macro', 'financial', 'international'].forEach(function(cat) {
    var el = document.getElementById('cards-' + cat);
    if (!el) return;
    new Sortable(el, {
      group:       'stories',
      animation:   200,
      ghostClass:  'card-ghost',
      chosenClass: 'card-chosen',
      dragClass:   'card-drag',
      handle:      '.drag-handle',
      onEnd: function(evt) {
        if (evt.from !== evt.to) {
          var newCat = evt.to.closest('.cat-section').dataset.category;
          evt.item.dataset.category      = newCat;
          evt.item.style.setProperty('--card-accent', ACCENTS[newCat] || '#2563eb');
          updateCounts();
        }
      }
    });
  });
}

/* ── Bootstrap ── */
document.addEventListener('DOMContentLoaded', function() {
  initSortable();
});
document.addEventListener('keydown', function(e) {
  if (e.key === 'Escape') closeModal();
});
/* Stop click-through on modal box */
document.getElementById('modal').addEventListener('click', function(e) {
  e.stopPropagation();
});
document.getElementById('pw-modal').addEventListener('click', function(e) {
  e.stopPropagation();
});

/* ── Editor password gate ── */
var EDITOR_HASH   = '865ed16d3bc1a45492524f4b4c050fff4f4f6875637cc46820c76ffbb646504c';
var PUBLISH_URL   = 'NETLIFY_FUNCTION_URL_PLACEHOLDER';
var editorSession = null;

async function _sha256(s) {
  var buf = await crypto.subtle.digest('SHA-256', new TextEncoder().encode(s));
  return Array.from(new Uint8Array(buf)).map(function(b){ return b.toString(16).padStart(2,'0'); }).join('');
}
function openPwModal() {
  document.getElementById('pw-input').value = '';
  document.getElementById('pw-error').style.display = 'none';
  document.getElementById('pw-overlay').classList.add('active');
  document.getElementById('pw-modal').classList.add('active');
  setTimeout(function(){ document.getElementById('pw-input').focus(); }, 60);
}
function closePwModal() {
  document.getElementById('pw-overlay').classList.remove('active');
  document.getElementById('pw-modal').classList.remove('active');
}
async function checkEditorPassword() {
  var input = document.getElementById('pw-input').value;
  var hash  = await _sha256(input);
  if (hash === EDITOR_HASH) {
    editorSession = input;
    closePwModal();
    enterEditMode();
  } else {
    document.getElementById('pw-error').style.display = 'block';
    document.getElementById('pw-input').select();
  }
}
function enterEditMode() { document.body.classList.add('edit-mode'); }
function exitEditMode() { document.body.classList.remove('edit-mode'); editorSession = null; }
async function publishChanges() {
  if (PUBLISH_URL === 'NETLIFY_FUNCTION_URL_PLACEHOLDER') {
    alert('Publish endpoint not configured yet.');
    return;
  }
  var btn = document.querySelector('.btn-publish');
  btn.textContent = 'Saving\u2026';
  btn.disabled = true;
  document.body.classList.remove('edit-mode');
  var snapshot = '<!DOCTYPE html>\n' + document.documentElement.outerHTML;
  document.body.classList.add('edit-mode');
  try {
    var res = await fetch(PUBLISH_URL, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ password: editorSession, html: snapshot })
    });
    if (res.ok) {
      btn.innerHTML = '&#10003; Published!';
      setTimeout(function(){ btn.innerHTML = '&#8679; Publish'; btn.disabled = false; }, 3000);
    } else { throw new Error('Server returned ' + res.status); }
  } catch(err) {
    alert('Publish failed: ' + err.message);
    btn.innerHTML = '&#8679; Publish';
    btn.disabled = false;
  }
}
</script>
<button id="btn-toggle-all" class="btn-toggle-all" onclick="toggleAll()">&#9660; Expand All</button>

<!-- ══ EDITOR LOCK BUTTON ══ -->
<button class="btn-editor-lock" onclick="openPwModal()" title="Editor login">&#128274;</button>

<!-- ══ EDIT MODE BAR ══ -->
<div class="edit-mode-bar">
  <span class="edit-mode-label">&#9998; Editor</span>
  <button class="btn-publish" onclick="publishChanges()">&#8679; Publish</button>
  <button class="btn-exit-edit" onclick="exitEditMode()">Exit</button>
</div>

<!-- ══ PASSWORD MODAL ══ -->
<div class="pw-overlay" id="pw-overlay" onclick="closePwModal()"></div>
<div class="pw-modal" id="pw-modal">
  <div class="pw-modal-head">
    <span>&#128274;&nbsp; Editor Login</span>
    <button class="modal-close" onclick="closePwModal()">&#215;</button>
  </div>
  <div class="pw-modal-body">
    <div class="form-group">
      <label class="form-label" for="pw-input">Password</label>
      <input id="pw-input" type="password" class="input-field" placeholder="Enter editor password"
             onkeydown="if(event.key===&#39;Enter&#39;) checkEditorPassword()"/>
      <p class="pw-error" id="pw-error">Incorrect password &mdash; please try again.</p>
    </div>
  </div>
  <div class="modal-foot">
    <button class="btn-cancel" onclick="closePwModal()">Cancel</button>
    <button class="btn-save" onclick="checkEditorPassword()">Unlock</button>
  </div>
</div>
</body>
</html>"""


def generate_html(grouped, source_results, total_raw, total_relevant):
    """Build v2 interactive HTML report using the template."""
    week_start    = ONE_WEEK_AGO.strftime("%b %d")
    week_end      = TODAY.strftime("%b %d, %Y")
    total_stories = len(grouped)
    active_sources = sum(1 for r in source_results if r["status"] == "ok")

    sections_html = _build_sections(grouped)

    cat_nav = "".join(
        f'<a href="#section-{cat}" class="cat-nav-pill">'
        f'{CATEGORY_META[cat]["icon"]}&nbsp;{CATEGORY_META[cat]["label"]}</a>'
        for cat in CAT_ORDER
    )

    html = get_html_template()
    for placeholder, value in [
        ("PLACEHOLDER_WEEK_START",     week_start),
        ("PLACEHOLDER_WEEK_END",       week_end),
        ("PLACEHOLDER_TOTAL_STORIES",  str(total_stories)),
        ("PLACEHOLDER_TOTAL_RELEVANT", str(total_relevant)),
        ("PLACEHOLDER_TOTAL_RAW",      str(total_raw)),
        ("PLACEHOLDER_ACTIVE_SOURCES", str(active_sources)),
        ("PLACEHOLDER_SECTIONS",       sections_html),
        ("PLACEHOLDER_CAT_NAV",        cat_nav),
        ("PLACEHOLDER_GENERATED_AT",   TODAY.strftime("%Y-%m-%d %H:%M UTC")),
    ]:
        html = html.replace(placeholder, value)

    return html


# ─────────────────────────────────────────
# Excel Source Status Report
# ─────────────────────────────────────────

def generate_excel(source_results, excel_path):
    """Write source status to a formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Source Status"

    # ── Header row ──
    headers = ["Source", "Type", "Status", "Articles This Week", "Detail / Reason"]
    hdr_fill = PatternFill("solid", fgColor="0A1628")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws.row_dimensions[1].height = 22

    # ── Status fills ──
    ok_fill   = PatternFill("solid", fgColor="D1FAE5")   # green
    fail_fill = PatternFill("solid", fgColor="FEE2E2")   # red
    skip_fill = PatternFill("solid", fgColor="FEF9C3")   # yellow

    ok_font   = Font(color="065F46", size=10)
    fail_font = Font(color="991B1B", size=10)
    skip_font = Font(color="78350F", size=10)

    row = 2

    def write_row(data_row, status_str, fill, font_style, detail):
        for col_idx, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = font_style
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 5))
            cell.border    = Border(
                left=thin_side, right=thin_side,
                top=Side(style="thin", color="E5E7EB"),
                bottom=Side(style="thin", color="E5E7EB"),
            )

    # Successful / failed sources from RSS_SOURCES
    for r in source_results:
        status = r["status"]
        if status == "ok":
            fill, fnt = ok_fill, ok_font
            status_str = "OK"
            detail     = r.get("detail", "")
            week_count = r.get("article_count", 0)
        else:
            fill, fnt = fail_fill, fail_font
            status_str = "Failed"
            detail     = r.get("error", "")
            week_count = 0

        ws.cell(row=row, column=1, value=r["name"]).fill    = fill
        ws.cell(row=row, column=2, value=r.get("type","")).fill = fill
        ws.cell(row=row, column=3, value=status_str).fill   = fill
        ws.cell(row=row, column=4, value=week_count).fill   = fill
        ws.cell(row=row, column=5, value=detail).fill       = fill

        for col_idx in range(1, 6):
            cell = ws.cell(row=row, column=col_idx)
            cell.font      = fnt
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 5))
            cell.border    = Border(
                left=thin_side, right=thin_side,
                top=Side(style="thin", color="E5E7EB"),
                bottom=Side(style="thin", color="E5E7EB"),
            )
        row += 1

    # Skipped / not automated sources
    for s in SKIPPED_SOURCES:
        detail = f"{s['reason']}  |  Suggestion: {s['suggestion']}"
        ws.cell(row=row, column=1, value=s["name"]).fill = skip_fill
        ws.cell(row=row, column=2, value="–").fill       = skip_fill
        ws.cell(row=row, column=3, value="Skipped").fill = skip_fill
        ws.cell(row=row, column=4, value=0).fill         = skip_fill
        ws.cell(row=row, column=5, value=detail).fill    = skip_fill

        for col_idx in range(1, 6):
            cell = ws.cell(row=row, column=col_idx)
            cell.font      = skip_font
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 5))
            cell.border    = Border(
                left=thin_side, right=thin_side,
                top=Side(style="thin", color="E5E7EB"),
                bottom=Side(style="thin", color="E5E7EB"),
            )
        row += 1

    # ── Column widths ──
    col_widths = [28, 10, 12, 20, 70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # ── Row heights for data rows ──
    for r_idx in range(2, row):
        ws.row_dimensions[r_idx].height = 18

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    wb.save(excel_path)


# ─────────────────────────────────────────
# Main
# ─────────────────────────────────────────

def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY environment variable is not set.")
        print("Set it with: set ANTHROPIC_API_KEY=sk-ant-...")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    # Force UTF-8 output on Windows
    if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
        sys.stdout = open(sys.stdout.fileno(), mode="w", encoding="utf-8", buffering=1)

    print("=" * 65)
    print("  MEXICO WEEKLY NEWS AGGREGATOR  v2")
    print(f"  Period: {ONE_WEEK_AGO.strftime('%Y-%m-%d')} to {TODAY.strftime('%Y-%m-%d')}")
    print("=" * 65)

    # ── Step 1: Fetch RSS feeds ──────────────────────────────────
    print("\n[1/4] Fetching RSS feeds...\n")
    all_articles   = []
    source_results = []

    for src in RSS_SOURCES:
        rss_display = src['rss'][:60] + "..." if len(src['rss']) > 60 else src['rss']
        print(f"  -> {src['name']} ({rss_display})")
        articles, error, detail, week_count = fetch_rss(src)
        if error:
            print(f"     FAILED: {error}")
            source_results.append({
                "name": src["name"], "type": src["type"],
                "status": "fail", "error": error,
            })
        else:
            print(f"     OK: {detail}")
            all_articles.extend(articles)
            source_results.append({
                "name": src["name"], "type": src["type"],
                "status": "ok", "detail": detail, "article_count": week_count,
            })

    total_raw = len(all_articles)
    print(f"\n  Total raw articles collected: {total_raw}")

    # ── Always write Excel source status ─────────────────────────
    print(f"\n  Writing source status Excel...")
    generate_excel(source_results, EXCEL_FILE)
    print(f"  Excel saved to: {EXCEL_FILE}")

    if total_raw == 0:
        print("\nNo articles fetched. Check your network or RSS URLs.")
        sys.exit(1)

    # ── Step 2: Filter relevant articles with Claude ─────────────
    print(f"\n[2/4] Filtering for Mexico relevance (Claude)...")
    relevant = filter_relevant(all_articles, client, batch_size=20)
    total_relevant = len(relevant)
    print(f"  Relevant articles: {total_relevant}/{total_raw}")

    if total_relevant == 0:
        print("\nNo relevant articles found after filtering.")
        html = generate_html([], source_results, total_raw, 0)
        with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"\nHTML (empty) saved to: {OUTPUT_FILE}")
        return

    # ── Step 3: Group & summarize with Claude ────────────────────
    print(f"\n[3/4] Grouping & summarizing stories (Claude)...")
    grouped = group_and_summarize(relevant, client)
    print(f"  Grouped into {len(grouped)} unique stories")

    # ── Step 4: Generate HTML ─────────────────────────────────────
    print(f"\n[4/4] Generating HTML report (v2)...")
    html = generate_html(grouped, source_results, total_raw, total_relevant)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n{'='*65}")
    print(f"  DONE")
    print(f"  HTML  -> {OUTPUT_FILE}")
    print(f"  Excel -> {EXCEL_FILE}")
    print(f"\n  Stories : {len(grouped)}")
    print(f"  Relevant: {total_relevant} / {total_raw} raw")
    ok_names   = [r['name'] for r in source_results if r['status'] == 'ok']
    fail_names = [r['name'] for r in source_results if r['status'] == 'fail']
    print(f"  Sources OK    : {', '.join(ok_names) or 'none'}")
    print(f"  Sources FAIL  : {', '.join(fail_names) or 'none'}")
    print("=" * 65)


if __name__ == "__main__":
    main()
